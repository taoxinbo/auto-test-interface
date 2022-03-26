# encoding=utf-8
# 2020-10-28
# TaoXB
# 此文件是接口自动化测试:提取报文标签里中的值

import os
import sys

# 当前文件所在目录都添加到sys.path中，系统可以找到需要引用的模块
sys.path.append(os.path.join(os.path.dirname(os.path.realpath(__file__)), ".."))
import requests
import re
import random
import string
import time
import datetime
from Action.Log import *
from Config.VarConfig import *
import xlrd
import traceback
import cx_Oracle
import pymysql
from openpyxl import load_workbook


# 获取表格的总行数
def get_row_num(excel_path, sheet):
    wb = load_workbook(excel_path)
    sheet = wb[sheet]
    rows = sheet.max_row
    return rows


# 获取表格的总列数
def get_clo_num(excel_path, sheet):
    wb = load_workbook(excel_path)
    sheet = wb[sheet]
    columns = sheet.max_column
    return columns


# 在XML中获取标签值为数字，每次运行都获取随机值相加，保证每次执行的值都唯一，然后替换之前旧的值。
def label_num(labelID, body):  # labelID是要替换的标签;body是需要处理的XML报文体。
    # temp = re.search(a, body)
    temp = re.search(r'<%s>(.*?)</%s>' % (labelID, labelID), body)
    # 判断标签值不为空，进行替换，保证每次执行的值都唯一
    if temp is not None:
        label_temp = int(temp.group(1)) + random.randint(1, 99999)
        return body.replace(temp.group(1), str(label_temp))
    # 判断标签值不存在的时候，返回最原始XML报文体body
    else:
        return body


# 在XML中获取标签值为‘数字’和‘字母’的组合，每次运行都获取随机值组合，保证每次执行的值都唯一，然后替换之前旧的值。
def label_str(labelID, body):  # labelID是要替换的标签;body是需要处理的XML报文体。
    # temp = re.search(a, body)
    temp = re.search(r'<%s>(.*?)</%s>' % (labelID, labelID), body)
    # 判断标签值不为空，进行替换，保证每次执行的值都唯一
    if temp is not None:
        label_temp = "".join([random.choice(list(string.digits)) for i in range(5)]) + "".join(
            [random.choice(list(string.ascii_uppercase)) for i in range(3)])
        return body.replace(temp.group(1), str(label_temp))
    # 判断标签值不存在的时候，返回最原始XML报文体body
    else:
        return body


# 在XML中获取标签值'年月日时分秒'并且可以'提前'或'延后'分钟数，每次运行都获取最新的'年月日时分秒'，保证每次执行的值都唯一，然后替换之前旧的值。
def label_YmdHMS(labelID, body, min):  # labelID是要替换的标签;body是需要处理的XML报文体;min延迟或提前的分钟数，min可以等于0。
    # temp = re.search(a, body)
    temp = re.search(r'<%s>(.*?)</%s>' % (labelID, labelID), body)
    # 判断标签值不为空，进行替换，保证每次执行的值都唯一
    if temp is not None:
        # label_temp = int(temp.group(1)) + random.randint(1, 99999)
        label_temp = (datetime.datetime.now() - datetime.timedelta(minutes=min)).strftime("%Y-%m-%d %H:%M:%S")
        return body.replace(temp.group(1), str(label_temp))
    # 判断标签值不存在的时候，返回最原始XML报文体body
    else:
        return body


# 在XML中获取标签值'年'，每次运行都随机获取一个年值，保证每次执行的值都唯一，然后替换之前旧的值。
def label_Y(labelID, body):  # labelID是要替换的标签;body是需要处理的XML报文体。
    # temp = re.search(a, body)
    temp = re.search(r'<%s>(.*?)</%s>' % (labelID, labelID), body)
    # 判断标签值不为空，进行替换，保证每次执行的值都唯一
    if temp is not None:
        label_temp = random.randint(2020, 9999)
        # label_temp = (datetime.datetime.now() - datetime.timedelta(minutes=min)).strftime("%H:%M:%S")
        return body.replace(temp.group(1), str(label_temp))
    # 判断标签值不存在的时候，返回最原始XML报文体body
    else:
        return body


# 在XML中获取标签值'年月日'并且可以'提前'或'延后'天数，每次运行都获取最新的'年月日'，保证每次执行的值都唯一，然后替换之前旧的值。
def label_Ymd(labelID, body, day):  # labelID是要替换的标签;body是需要处理的XML报文体;day延迟或提前的天数，day可以等于0。
    # temp = re.search(a, body)
    temp = re.search(r'<%s>(.*?)</%s>' % (labelID, labelID), body)
    # 判断标签值不为空，进行替换，保证每次执行的值都唯一
    if temp is not None:
        # label_temp = int(temp.group(1)) + random.randint(1, 99999)
        label_temp = (datetime.datetime.now() - datetime.timedelta(days=day)).strftime("%Y-%m-%d")
        return body.replace(temp.group(1), str(label_temp))
    # 判断标签值不存在的时候，返回最原始XML报文体body
    else:
        return body


# 在XML中获取标签值'时分秒'并且可以'提前'或'延后'分钟数，每次运行都获取最新的'时分秒'，保证每次执行的值都唯一，然后替换之前旧的值。
def label_HMS(labelID, body, min):  # labelID是要替换的标签;body是需要处理的XML报文体;min延迟或提前的分钟数，min可以等于0。
    # temp = re.search(a, body)
    temp = re.search(r'<%s>(.*?)</%s>' % (labelID, labelID), body)
    # 判断标签值不为空，进行替换，保证每次执行的值都唯一
    if temp is not None:
        # label_temp = int(temp.group(1)) + random.randint(1, 99999)
        label_temp = (datetime.datetime.now() - datetime.timedelta(minutes=min)).strftime("%H:%M:%S")
        return body.replace(temp.group(1), str(label_temp))
    # 判断标签值不存在的时候，返回最原始XML报文体body
    else:
        return body


# 在XML中获取标签值为字符串，每次运行都获取当前的'年月日时分秒'相加，保证每次执行的值都唯一。
def label_time(labelID, body):  # labelID是要替换的标签;body是需要处理的XML报文体。
    # temp = re.search(a, body)
    temp = re.search(r'<%s>(.*?)</%s>' % (labelID, labelID), body)
    # 判断标签值不为空，进行替换，保证每次执行的值都唯一
    if temp is not None:
        t = time.strftime("%Y%m%d%H%M%S")
        return re.sub(r"(?<=<%s>)(\w+)(?=</%s>)" % (labelID, labelID), r"\1_" + t, body)
    # 判断标签值不存在的时候，返回最原始XML报文体body
    else:
        return body


# 读取Excel
def read_file(file_path):
    try:
        data = xlrd.open_workbook(file_path, formatting_info=True)
        return data
    except Exception as e:
        print(str(e))


# 读取某个目录下的excel文件中的每一格内容
def excel(file_path, sheets):
    try:
        # wb = xlrd.open_workbook(sys.path[0] +'\\20201214.xls')# sys.path[0]为要获取当前路径，20201214.xls为要打开的excel文件
        wb = read_file(file_path)  # 调用'读取Excel文件位置'的函数
        sheet = wb.sheet_by_name(sheets)  # 通过excel表格名称(rank)获取工作表
        dat = []  # 创建空list
        for a in range(sheet.nrows):  # 循环读取表格中的行数
            for b in range(sheet.ncols):  # 循环读取表格中的列数
                if isinstance(sheet.cell_value(a, b), (int, float)):
                    data = int(sheet.cell_value(a, b))
                else:
                    data = sheet.cell_value(a, b)  # 因为表内可能存在多列数据，0代表第一列数据，1代表第二列，以此类推
                dat.append(data)
        return dat
    except Exception as e:
        print(str(e))


# 循环读取列表中的值,按照一行显示n个值换行
def L1(a, n):  # a变量传入
    count = 0
    for i in a:  # for 语句循环 ，输出1~100的数字
        if len(str(i)) == 3:
            print(str(i) + ' ', end=" ")  # end 表示输出i 之后一个空间的表示,如当i= 5 ,end 定义为end = "," 时，输出为" 5,"(实际输出没有"")
        elif len(str(i)) == 2 and isinstance(i, (int, float)):
            print(2 * ' ' + str(i) + ' ', end=" ")  # end 表示输出i 之后一个空间的表示,如当i= 5 ,end 定义为end = "," 时，输出为" 5,"(实际输出没有"")
        else:
            print(i, end=" ")
        count += 1  # count 开始计次
        if count % n == 0:
            print(end="\n")  # 每行第5个数开始换行


# 获取当前的日期
def get_current_date():
    time_tup = time.localtime()
    current_date = str(time_tup.tm_year) + "年" + \
                   str(time_tup.tm_mon) + "月" + str(time_tup.tm_mday) + "日"
    return current_date


# 创建文件夹
def make_dir(dir_path):
    if not os.path.exists(dir_path):
        try:
            os.mkdir(dir_path)
        except Exception as e:
            print("创建%s目录失败" % dir_path)
            raise e


# 创建带有年月日的文件夹，'default_dir_path = None’的意思：在函数中对于参数 可以传 也可以 不传。
def make_current_date_dir(default_dir_path=None):
    if default_dir_path is None:
        dir_path = get_current_date()
    else:
        dir_path = os.path.join(default_dir_path, get_current_date())
    if not os.path.exists(dir_path):
        try:
            os.mkdir(dir_path)
        except Exception as e:
            print("创建%s目录失败" % dir_path)
            raise e
    return dir_path


# 返回当前SQL有多少数据，针对Oracle数据库
def Oracle_DB(ora_ip, ora_sid, ora_user, ora_pwd, sql1):
    # 创建数据库连接,配置监听并连接
    # cx_Oracle.connect('username','pwd','ora的tns信息')
    # oracle数据库的tns信息，从tnsnames.ora中找到plsql可用的配置项，将该配置项直接拷贝过来即可
    try:
        ora_tns = cx_Oracle.makedsn(ora_ip, 1521, ora_sid)
        conn = cx_Oracle.connect(ora_user, ora_pwd, ora_tns)
        # 操作游标
        cursor = conn.cursor()
        # 执行SQL，返回执行SQL的值
        cursor.execute(sql1)
        # number = cursor.fetchone()
        number = cursor.fetchall()
        # print(number)
        # print(len(number))
        # print(type(len(number)))
        # 关闭游标
        cursor.close()
        # 提交事务
        conn.commit()
        # 关闭数据库连接
        conn.close()
        return len(number)
    except (cx_Oracle.DatabaseError) as err:
        if 'ORA-01017' in str(err):
            print('数据库登陆失败', err)
            logging.debug("数据库登陆失败" + str(err))
            return ("数据库登陆失败")
        elif 'ORA-00942' in str(err):
            print('表或视图不存在', err)
            logging.debug("表或视图不存在" + str(err))
            return ("表或视图不存在")
    except Exception as err:
        # 发生其他异常时，打印异常堆栈信息
        # return ('执行SQL出现未知错误')
        print('执行SQL出现未知错误：', err)
        logging.debug("执行SQL出现未知错误：" + str(err))
        return ('执行SQL出现未知错误')


# 返回当前SQL有多少数据，针对mysql数据库
def MySQL_DB(my_host, my_port, my_user, my_passwd, my_db, sql1):
    try:
        connect = pymysql.Connect(
            host=my_host,
            port=my_port,
            user=my_user,
            passwd=my_passwd,
            db=my_db,
            charset='utf8'
        )
        cursor = connect.cursor()
        # 执行SQL，返回执行SQL的值
        cursor.execute(sql1)
        # number = cursor.fetchone()
        number = cursor.fetchall()
        # print(number)
        # print(len(number))
        # print(type(len(number)))
        # 关闭游标
        cursor.close()
        # 提交事务
        connect.commit()
        # 关闭数据库连接
        connect.close()
        return len(number)
    except (pymysql.err.ProgrammingError) as err:
        print("表或视图不存在", err)
        logging.debug("表或视图不存在" + str(err))
        return ("表或视图不存在")
    except (pymysql.err.OperationalError) as err:
        print("数据库登陆失败", err)
        logging.debug("数据库登陆失败" + str(err))
        return ("数据库登陆失败")
    except (pymysql.err.InternalError) as err:
        print("数据库不存在", err)
        logging.debug("数据库不存在" + str(err))
        return ("数据库不存在")
    except Exception as err:
        # 发生其他异常时，打印异常堆栈信息
        # return ('执行SQL出现未知错误')
        print('执行SQL出现未知错误：', err)
        logging.debug("执行SQL出现未知错误：" + str(err))
        return ('执行SQL出现未知错误')


if __name__ == '__main__':
    excel_path = baseDir + "/ExcelData/" + "Public1.xlsx"
    sheet = 'Public'

    print(get_clo_num(excel_path, sheet))
    print(type(get_clo_num(excel_path, sheet)))
    print(get_row_num(excel_path, sheet))

    # a=Oracle_DB('10.60.44.92','atsdb','jats033','jats033',"select * from T_SE_COUNTERPARTIES where name like '%接口自动化测试%'")
    # a = Oracle_DB('10.60.44.92','atsdb','imp','test1',"select * from T_MI_PAYMENTS where 1SRCBATCHNO like 'txbPMSQ01%'")
    # print(a)
    # if a==None:
    #     print(1)
    # b = MySQL_DB('10.60.44.221', 3306, 'root', "Mysql@123", 'jats2610reg',"select * from T_SE_COUNTERPARTIES where name like '%接口自动化测试%'")
    # print(b)
    # if b == None:
    #     print(1)
    # #获取根目录下的文件夹/XML_request/
    # baseDir1 = baseDir + "/XML_request/"
    # # 读取XML_request文件中的内容
    # a='body_1908'
    # with open(baseDir1 + a + '.xml', encoding='utf-8') as fp:
    #     body1 = fp.read()
    # logging.info("打开报文体：%s"%a)
    #
    # body2=label_num('TransSeqID', body1)
    # logging.info('标签%s替换成功'%'TransSeqID')
    #
    # if a in []:
    #     pass
    # else:
    #     body2=label_str('ReqSeqID',body2)
    # logging.info('标签%s替换成功' % 'ReqSeqID')
    #
    # body2=label_YmdHMS('TransTime',body2,0)
    # logging.info('标签%s替换成功' % 'TransTime')
    # #logging.info("="*80)
    #
    # body2=label_YmdHMS('BgnDate', body2, 10)
    # body2=label_Ymd('BookDate', body2, 0)
    # body2=label_Ymd('PayDate', body2, 0)
    # body2=label_HMS('PayTime', body2, 0)
    #
    # print('body1的值:\n%s' % body1)
    # #logging.info('body1的值:\n%s'%body1)
    #
    # print('body2的值:\n%s' % body2)
    # #logging.info('body2的值:\n%s' % body2)

    # wb = load_workbook(baseDir + "/ExcelData/" + "Public1.xlsx")
    # sheet = wb['Public']
    # print(sheet.max_row)
