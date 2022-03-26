# -*- coding: utf-8 -*-
# @Time : 2021/8/22 9:40
# @Author : chengguo
# @File : read_excel.py
# @脚本功能描述：
from openpyxl import load_workbook
from Config.VarConfig import *
from Action.extract import *


class DoExcsl:

    def __init__(self, sheet, excel_path, case_id):
        """
        :param write_baseAdd: excel表中对应的sheet
        :param excel_path: 测试用例路径
        :param case_id: 想要执行具体的用例参数有两个（all全部， caseid指定测试用例）
        """
        self.excel_path = excel_path
        self.sheet = sheet
        self.case_id = case_id

    def read_case(self):
        """
        :return:
        """
        wb = load_workbook(self.excel_path)
        sheet = wb[self.sheet]
        text_data = []
        for i in range(2, sheet.max_row + 1):
            sub_data = {}
            sub_data["Case_id"] = sheet.cell(i, 1).value
            sub_data["Title"] = sheet.cell(i, 2).value
            sub_data["APIName"] = sheet.cell(i, 3).value
            sub_data["Url"] = sheet.cell(i, 4).value
            sub_data["method"] = sheet.cell(i, 5).value
            sub_data["is_run"] = sheet.cell(i, 6).value
            sub_data["Params"] = sheet.cell(i, 7).value
            sub_data["ExpectedResult"] = sheet.cell(i, 8).value
            sub_data["DBType"] = sheet.cell(i, 9).value
            sub_data["Chick_sql"] = sheet.cell(i, 10).value
            sub_data["Result"] = sheet.cell(i, 11).value
            sub_data["error_msg"] = sheet.cell(i, 12).value
            text_data.append(sub_data)
        final_data = []
        if self.case_id == 'all':
            final_data = text_data
        else:
            for i in self.case_id:
                final_data.append(text_data[int(i) - 1])
        return final_data

    def write_excel(self, row, column, values):
        """
        :param row: 行
        :param column: 列
        :param values: 值
        :return:
        """
        try:
            wb = load_workbook(self.excel_path)
            sheet = wb[self.sheet]
            sheet.cell(row, column).value = values
            wb.save(self.excel_path)
            # print('保存成功')
        except KeyError as e:
            print("您的输入有误，报错为：{}".format(e))
            exit()


if __name__ == '__main__':
    # data = DoExcsl(case_id='all', sheet='Public',
    #                excel_path=r'C:\Users\F\Desktop\ATS_Interface_Auto_Test_Public\ExcelData\Public.xlsx').read_case()
    # print([tuple(i.values()) for i in data])

    excel_path1 = baseDir + "/ExcelData/" + "Public.xlsx"
    sheet1 = 'Push_case'

    data = DoExcsl(case_id='all', sheet=sheet1, excel_path=excel_path1).read_case()
    read_excel = DoExcsl(case_id='all', sheet=sheet1, excel_path=excel_path1)
    print(data)

    # sheet = wb['Public']
    # print(sheet.max_row)
    # print(sheet.max_column)

    # print(get_row_num(excel_path, sheet))
    for j in range(2, get_row_num(excel_path1, sheet1) + 1):
        # for j in range(2, 4):
        # print(j)
        read_excel.write_excel(j, test_res_col, '')
        read_excel.write_excel(j, test_inf_col, '')
